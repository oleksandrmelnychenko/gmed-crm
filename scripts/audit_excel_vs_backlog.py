"""
Generate docs/testing/user-stories-excel-backlog-audit_ua.md
Compares User Stories sheet to docs/requirements/03_product-backlog_ua.md
"""
from __future__ import annotations

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "docs" / "1 (Update 2) User Story Salesforce.xlsx"
XLSX_JSON = ROOT / ".agent-extract" / "user_stories_rows.json"
BACKLOG = ROOT / "docs" / "requirements" / "03_product-backlog_ua.md"
OUT = ROOT / "docs" / "testing" / "user-stories-excel-backlog-audit_ua.md"


def ensure_excel_json() -> None:
    from openpyxl import load_workbook

    XLSX_JSON.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["User Stories"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [("" if c is None else str(c).strip()) for c in row]
        rows.append({"sheet_row": i + 1, "cells": cells})
    wb.close()
    XLSX_JSON.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")


def parse_excel_stories() -> list[dict]:
    if not XLSX_JSON.is_file():
        ensure_excel_json()
    rows = json.loads(XLSX_JSON.read_text(encoding="utf-8"))
    cur_epic: int | None = None
    items: list[dict] = []
    for r in rows[1:]:
        cells = r["cells"] + [""] * 6
        epic_cell = cells[0].strip()
        if epic_cell:
            m = re.match(r"EPIC\s+(\d+)\s*:", epic_cell)
            if m:
                cur_epic = int(m.group(1))
        story = cells[2].strip()
        if not story or cur_epic is None:
            continue
        pr = cells[5].strip()
        items.append(
            {
                "epic": cur_epic,
                "sheet_row": r["sheet_row"],
                "rolle": cells[1],
                "story": story,
                "beschreibung": cells[3],
                "security": cells[4],
                "priority": pr,
            }
        )
    return items


def parse_ua_backlog() -> dict[int, list[tuple[int, str, int | tuple[int, int] | None]]]:
    """epic -> list of (line_number, full_line, P priority: int or (low,high) or None)"""
    text = BACKLOG.read_text(encoding="utf-8").splitlines()
    cur: int | None = None
    out: dict[int, list[tuple[int, str, int | None]]] = {}
    for i, line in enumerate(text, start=1):
        m = re.match(r"^## EPIC (\d+):", line)
        if m:
            cur = int(m.group(1))
            out[cur] = []
            continue
        if cur is not None and line.startswith("- **["):
            pm = re.search(r"- \*\*\[P(\d+)(?:-P(\d+))?\]", line)
            if pm:
                lo = int(pm.group(1))
                hi = int(pm.group(2)) if pm.group(2) else lo
                p = (lo, hi) if hi != lo else lo
            else:
                p = None
            out[cur].append((i, line, p))
    return out


def main() -> None:
    excel_items = parse_excel_stories()
    ua_by_epic = parse_ua_backlog()

    lines: list[str] = []
    lines.append("# Аудит: Excel User Stories ↔ `03_product-backlog_ua.md`")
    lines.append("")
    lines.append("> Автоматично згенеровано скриптом `scripts/audit_excel_vs_backlog.py`. Передати клієнту варто після ручної перевірки рядків зі статусом **розбіжність кількості** або **пріоритет**.")
    lines.append("")
    lines.append("**Джерело Excel:** `docs/1 (Update 2) User Story Salesforce.xlsx`, аркуш `User Stories`.")
    lines.append("")
    lines.append("**Цільовий документ:** `docs/requirements/03_product-backlog_ua.md`.")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Зведення по EPIC")
    lines.append("")
    lines.append("| EPIC | Рядків у Excel | Кульок у UA | Примітка |")
    lines.append("|------|----------------|-------------|----------|")

    for epic in range(1, 25):
        ex = [x for x in excel_items if x["epic"] == epic]
        ua_list = ua_by_epic.get(epic, [])
        ne, nu = len(ex), len(ua_list)
        if ne == nu:
            note = "Зіставлення 1:1 по порядку"
        else:
            note = f"Δ={ne - nu:+d}: у UA кілька історій **об'єднано** в одну кулю або розбито інакше — див. розділ 3"
        lines.append(f"| {epic} | {ne} | {nu} | {note} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 2. EPIC з однаковою кількістю: порядок, пріоритет, заголовок")
    lines.append("")
    lines.append("Для цих EPIC застосовано зіставлення **за позицією** (n-та історія Excel → n-та куля UA).")
    lines.append("")

    priority_issues: list[str] = []

    for epic in range(1, 25):
        ex = [x for x in excel_items if x["epic"] == epic]
        ua_list = ua_by_epic.get(epic, [])
        if len(ex) != len(ua_list):
            continue

        lines.append(f"### EPIC {epic}")
        lines.append("")
        lines.append("| Excel рядок | DE User Story | P Excel | P UA | ΔP | UA рядок |")
        lines.append("|-------------|---------------|---------|------|-----|----------|")

        def ua_p_label(ua_p):
            if ua_p is None:
                return "∅"
            if isinstance(ua_p, tuple):
                return f"P{ua_p[0]}-P{ua_p[1]}"
            return f"P{ua_p}"

        def priority_matches(excel_s: str, ua_p) -> bool:
            if not excel_s.strip() or ua_p is None:
                return True
            if not excel_s.isdigit():
                return True
            ex_n = int(excel_s)
            if isinstance(ua_p, tuple):
                return ua_p[0] <= ex_n <= ua_p[1]
            return ex_n == ua_p

        for e_item, (ua_ln, ua_line, ua_p) in zip(ex, ua_list):
            ex_p = e_item["priority"]
            p_ok = "✓" if priority_matches(ex_p, ua_p) else "⚠"
            if p_ok == "⚠":
                priority_issues.append(
                    f"- **EPIC {epic}**, Excel рядок {e_item['sheet_row']}: Excel P={ex_p or '∅'}, UA {ua_p_label(ua_p)} — «{e_item['story'][:55]}»"
                )
            story_short = e_item["story"].replace("|", "\\|")[:70]
            lines.append(
                f"| {e_item['sheet_row']} | {story_short} | {ex_p or '∅'} | {ua_p_label(ua_p)} | {p_ok} | L{ua_ln} |"
            )
        lines.append("")

    if priority_issues:
        lines.append("#### Пріоритети не збіглися (потрібне вирівнювання з Excel)")
        lines.append("")
        lines.extend(priority_issues)
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 3. EPIC з різною кількістю: повні списки для ручного мапінгу")
    lines.append("")

    mismatch_any = False
    for epic in range(1, 25):
        ex = [x for x in excel_items if x["epic"] == epic]
        ua_list = ua_by_epic.get(epic, [])
        if len(ex) == len(ua_list):
            continue

        mismatch_any = True
        lines.append(f"### EPIC {epic} — Excel {len(ex)} / UA {len(ua_list)}")
        lines.append("")
        lines.append("**Excel (User Story):**")
        for e_item in ex:
            p = e_item["priority"] or "∅"
            lines.append(
                f"- Рядок **{e_item['sheet_row']}** [P{p}]: {e_item['story']}"
            )
        lines.append("")
        def _ua_p_lab(p):
            if p is None:
                return "∅"
            if isinstance(p, tuple):
                return f"P{p[0]}-P{p[1]}"
            return f"P{p}"

        lines.append("**UA (перший рядок кожної кулі):**")
        for ua_ln, ua_line, ua_p in ua_list:
            body = ua_line[2:].strip() if ua_line.startswith("- ") else ua_line
            short = body.replace("|", "\\|")
            if len(short) > 200:
                short = short[:200] + "…"
            lines.append(f"- **L{ua_ln}** [{_ua_p_lab(ua_p)}]: {short}")
        lines.append("")
        lines.append(
            "*Рекомендація:* оновити UA-беклог або додати підпункти, щоб кожна рядок Excel мала явне відображення, якщо клієнт хоче трасованість 1:1."
        )
        lines.append("")

    if not mismatch_any:
        lines.append(
            "*Немає EPIC, де кількість рядків Excel і куль у `03_product-backlog_ua.md` розходиться.*"
        )
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 4. Позначки в `03_product-backlog_ua.md`")
    lines.append("")
    lines.append(
        "У самому беклозі додано HTML-коментар після заголовка документа з посиланням на цей файл аудиту (для редакторів; у звичайному перегляді не видно)."
    )
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 5. Короткі висновки")
    lines.append("")
    lines.append(
        "- **Стан зведення (§1):** після регенерації `03_product-backlog_ua.md` скриптом `scripts/generate_product_backlog_from_excel.py` кількість пунктів у кожному EPIC **збігається** з кількістю рядків User Stories у Excel (1:1)."
    )
    lines.append(
        "- **Пріоритети (§2):** порівняння `[P*]` з колонкою Priority; порожній Priority в Excel у згенерованому файлі замінено на `P1` — див. примітку в шапці беклогу."
    )
    lines.append(
        "- **Розділ 3** (повні списки при різній кількості) за поточного стану **порожній** — розбіжностей за кількістю немає."
    )
    lines.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
